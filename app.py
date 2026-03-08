from flask import Flask, render_template, request, send_file, redirect
import csv
import os
import openpyxl

app = Flask(__name__)

EXCEL_FILE = "daily_status.xlsx"


def save_to_excel(name, date, project, work_done, blockers, plan):

    file_exists = os.path.isfile(EXCEL_FILE)

    if file_exists:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Project", "Work Done", "Blockers", "Plan"])

    ws.append([name, date, project, work_done, blockers, plan])
    wb.save(EXCEL_FILE)


def load_reports_from_excel():
    if not os.path.isfile(EXCEL_FILE):
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    reports = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row and len(row) >= 5:  # Ensure at least 5 columns
            reports.append({
                "name": row[0] or "",
                "date": row[1] or "",
                "project": row[2] if len(row) > 5 else "",  # Project column if exists
                "work_done": row[2] if len(row) <= 5 else row[3] or "",
                "blockers": row[3] if len(row) <= 5 else row[4] or "",
                "plan": row[4] if len(row) <= 5 else row[5] or ""
            })
    return reports


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]
        role = request.form["role"]
        # For now, just print the data (in a real app, you'd authenticate)
        print(f"Login attempt: Name={name}, Email={email}, Role={role}")
        # Redirect to index after login
        return redirect("/index")
    return render_template("login.html")


@app.route("/index", methods=["GET", "POST"])
def dashboard():

    report = None

    if request.method == "POST":

        name = request.form["name"]
        date = request.form["date"]
        project = request.form["project"]
        status = request.form["status"]

        lines = status.split("\n")

        work_done = lines[0] if len(lines) > 0 else ""
        blockers = lines[1] if len(lines) > 1 else ""
        plan = lines[2] if len(lines) > 2 else ""

        save_to_excel(name, date, project, work_done, blockers, plan)

        report = {
            "name": name,
            "date": date,
            "project": project,
            "work_done": work_done,
            "blockers": blockers,
            "plan": plan
        }

    return render_template("index.html", report=report)


@app.route("/reports")
def reports():
    all_reports = load_reports_from_excel()
    total_reports = len(all_reports)
    reports_with_projects = len([r for r in all_reports if r.get('project', '').strip()])
    contributors = len([r for r in all_reports if r.get('name', '').strip()])
    return render_template("reports.html", reports=all_reports, total_reports=total_reports, reports_with_projects=reports_with_projects, contributors=contributors)


@app.route("/download")
def download():
    if os.path.isfile(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    else:
        return "No Excel file found", 404


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
